from pygeocoder import Geocoder
import openpyxl

workbook = openpyxl.load_workbook('Companies.xlsx')
sheet = workbook.get_sheet_by_name('Sheet1')

states = ['Michigan', 'Ohio', 'Pennsylvania', 'Kentuky', 'West Virginia']

for i in range(2018,7204,1):
    curCellValue = sheet.cell(row=i,column=1).value
    if  curCellValue == None:

        if curCellValue == result:
            print(i, result, "Yes Andrew")
            sheet.cell(row=i,column=1).value = "Andrew Pierpoint"
        else:
            result = Geocoder.geocode(sheet.cell(row=i, column=2).value+" HQ").state
           
            if any([x in result for x in states]):
                print(i, result, "Yes Andrew")
                sheet.cell(row=i,column=1).value = "Andrew Pierpoint"
            elif result == 'New York':
                sheet.cell(row=i,column=1).value = "TJ or Andrew Pierpoint"
                print(i, result, "Maybe Andrew")
            else:
                print (i, result, "Not Andrew")


    else:
        sheet.cell(row=i,column=12).value = "Didn't check"
        print (i, "Didn't Check")


workbook.save('UpdatedList2.xlsx')
#2018

#result = Geocoder.geocode("BANCO DE LA NACION HQ")



#print result

# The result for this is 6301 Fitch Path, New Albany, OH 43054, USA

# We can pull the city and/or state and then based on whose territory
# it is I can assign the BDR to that company on an excel file