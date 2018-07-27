# This script will place the state of the customer's HQ in the 13th column and place Andrew Pierpoint in the first column
# if it it his territory. The state the customer belongs to will not be placed if the BDR has already input their name.
# This is to save on querying the Google Maps API (limit ~2500 a day)

from pygeocoder import Geocoder
import openpyxl, time

workbook = openpyxl.load_workbook('Companies.xlsx') # Workbook must be named "Companies.xlsx"
sheet = workbook.get_sheet_by_name('Sheet1') # The sheet must be named "Sheet1"

andrewsStates = ['Michigan', 'Ohio', 'Pennsylvania', 'Kentuky', 'West Virginia']
result = ""

# Irerate over the entire list of companies
for i in range(2018,7204,1):
    curCellValue = sheet.cell(row=i,column=1).value
    
    #If the BDR value is empty then proceed
    if  curCellValue == None:
        try:
            
            # If the company is the same as the iteration prior then input the same values
            if curCellValue == result:
                print(i, result, "Same as above")
                sheet.cell(row=i,column=1).value = sheet.cell(row=i-1,column=1).value
            	sheet.cell(row=i,column=13).value = result

           	# If it's Andrew's territory then input his name under the BDR column
           	# This will always input the state at the end as long as column1 is empty
            else:
            	result = Geocoder.geocode(sheet.cell(row=i, column=2).value+" HQ").state
            	sheet.cell(row=i,column=13).value = result

                if any([x in result for x in andrewsStates]):
                    print(i, result, "Yes Andrew")
                    sheet.cell(row=i,column=1).value = "Andrew Pierpoint"
                
                elif result == 'New York':
                    sheet.cell(row=i,column=1).value = "TJ or Andrew Pierpoint"
                    print(i, result, "Maybe Andrew")
                
                else:
                    print (i, result, "Not Andrew")

        # Exceptions in this process come from no results returned (mutliple results will choose the top one)
        except:
            print (i, "No results")

    #If a BDR has already input their information then don't check
    else:
        sheet.cell(row=i,column=12).value = "Didn't check"
        print (i, "Didn't Check")

# Name the new workbook "UpdatedList" + the time this finished
_time = time.strftime("%Y%m%d_%H%M")
workbook.save('UpdatedList'+_time+'.xlsx')